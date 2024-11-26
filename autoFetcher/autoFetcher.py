import re
import time
import random
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import logging
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook

logging.basicConfig(filename='automation.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

options = webdriver.ChromeOptions()
options.add_experimental_option('detach', True)
options.add_argument(r'user-data-dir=C:\Users\aadit\AppData\Local\Google\Chrome\User Data')
options.add_argument("start-maximized")
options.add_argument("disable-infobars")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("window-size=1920x1080")

driver = webdriver.Chrome(options=options)


def wait_for(by, value, condition=EC.element_to_be_clickable):
    try:
        return WebDriverWait(driver, 10).until(condition((by, value)))
    except Exception as exc:
        logging.error(f"Error waiting for element {by} - {value}: {exc}")
        return None  # Return None instead of raising an exception


file_path = r"D:\Aadit\PESL\Book1.xlsx"
book = load_workbook(file_path)
pasted_sheet = book['Details']
pasted_rows = pasted_sheet.max_row
file = pd.read_excel(file_path, sheet_name='Connections')
links = file['URL'].dropna()[pasted_rows-1:].tolist()
data_list = []
desired_order = ['URL', 'Full Name', 'Company', 'Position', 'City', 'Phone', 'Email', 'Birthday', 'Connected', 'Address', 'Website', 'IM']

count = 0
for link in links:
    time.sleep(random.randint(2, 5))
    count += 1
    try:
        row = file[file['URL'] == link].iloc[0]
        data = {
            'URL': link,
            'Full Name': row['First Name'] + ' ' + row['Last Name'],
            'Company': row['Company'],
            'Position': row['Position'],
        }

        driver.get(link)

        contact_info = wait_for(By.ID, 'top-card-text-details-contact-info')
        contact_info.click()

        city_element = wait_for(By.CSS_SELECTOR, '.text-body-small.inline.t-black--light.break-words')
        data['City'] = city_element.text.split(',')[0] if city_element else 'N/A'

        details = wait_for(By.CSS_SELECTOR, '.pv-contact-info__contact-type', EC.presence_of_all_elements_located)
        if details:
            details = details[1:]  # Skip the first element
            for d in details:
                try:
                    tag = d.find_element(By.XPATH, './h3').text
                    val = d.find_element(By.CSS_SELECTOR, '.vaBaUoyvRPVHsMdwsdganXqklSMryPUp.link-without-visited-state.t-14').text

                    data[tag] = val

                except NoSuchElementException as e:
                    logging.error(f"Error finding details for {link} - {tag}: {e}")
                    data[tag] = 'N/A'  # Default value for missing data

        print(data)
        data_list.append(data)

        close_btn = wait_for(By.CSS_SELECTOR,'.artdeco-button.artdeco-button--circle.artdeco-button--muted.artdeco-button--2.artdeco-button--tertiary.ember-view.artdeco-modal__dismiss')
        close_btn.click()

    except Exception as e:
        logging.error(f"Error processing link {link}: {e}")

    if count % 50 == 0:
        all_keys = set().union(*(d.keys() for d in data_list))

        df = pd.DataFrame([{k: d.get(k, '') for k in all_keys} for d in data_list])[desired_order]

        file_path = r"D:\Aadit\PESL\Book1.xlsx"
        book = load_workbook(file_path)
        startrow = book['Details'].max_row

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name='Details', startrow=startrow)
        data_list = []

        if count==100:
            break